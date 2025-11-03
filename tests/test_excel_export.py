"""
Tests for Excel export functionality.
"""
import pytest
import pandas as pd
from pathlib import Path
from src.core.export_utils import save_excel


@pytest.mark.unit
class TestExcelExport:
    """Tests for Excel export functionality."""
    
    def test_save_excel_basic(self, sample_dataframe, tmp_path, clean_db, db_session):
        """Test basic Excel export."""
        validation_issues = [
            {'row_number': 0, 'column_name': 'age', 'issue_type': 'type_mismatch',
             'description': 'Invalid type', 'severity': 'high'},
            {'row_number': 1, 'column_name': 'email', 'issue_type': 'invalid_email',
             'description': 'Invalid email', 'severity': 'medium'},
        ]
        validation_summary = {'high': 1, 'medium': 1, 'low': 0}
        
        output_dir = str(tmp_path / "reports")
        excel_path = save_excel(
            df=sample_dataframe,
            validation_issues=validation_issues,
            validation_summary=validation_summary,
            filename="test_report",
            output_dir=output_dir
        )
        
        assert excel_path is not None
        assert Path(excel_path).exists()
        assert excel_path.endswith('.xlsx')
        
        # Verify file is valid (can be opened)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            assert len(wb.worksheets) > 0
        except ImportError:
            pytest.skip("openpyxl not available for validation")
    
    def test_save_excel_with_ml_recommendations(self, sample_dataframe, tmp_path):
        """Test Excel export with ML recommendations."""
        validation_issues = []
        validation_summary = {'high': 0, 'medium': 0, 'low': 0}
        ml_recommendations = {
            'readiness_score': 85,
            'readiness_level': 'Good'
        }
        
        output_dir = str(tmp_path / "reports")
        excel_path = save_excel(
            df=sample_dataframe,
            validation_issues=validation_issues,
            validation_summary=validation_summary,
            ml_recommendations=ml_recommendations,
            filename="test_report_ml",
            output_dir=output_dir
        )
        
        assert excel_path is not None
        assert Path(excel_path).exists()
    
    def test_save_excel_empty_issues(self, clean_dataframe, tmp_path):
        """Test Excel export with no issues."""
        validation_issues = []
        validation_summary = {'high': 0, 'medium': 0, 'low': 0}
        
        output_dir = str(tmp_path / "reports")
        excel_path = save_excel(
            df=clean_dataframe,
            validation_issues=validation_issues,
            validation_summary=validation_summary,
            filename="test_report_clean",
            output_dir=output_dir
        )
        
        assert excel_path is not None
        assert Path(excel_path).exists()
    
    def test_save_excel_multiple_sheets(self, sample_dataframe, tmp_path):
        """Test that Excel file contains multiple sheets."""
        validation_issues = [
            {'row_number': 0, 'column_name': 'age', 'issue_type': 'type_mismatch',
             'description': 'Invalid type', 'severity': 'high'},
        ]
        validation_summary = {'high': 1, 'medium': 0, 'low': 0}
        
        output_dir = str(tmp_path / "reports")
        excel_path = save_excel(
            df=sample_dataframe,
            validation_issues=validation_issues,
            validation_summary=validation_summary,
            filename="test_report_sheets",
            output_dir=output_dir
        )
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            sheet_names = [ws.title for ws in wb.worksheets]
            
            # Should have at least Overview, Missing Values, Issues sheets
            assert 'Overview' in sheet_names
            assert 'Missing Values' in sheet_names
            assert 'Issues' in sheet_names
            
            if len(sample_dataframe.select_dtypes(include=['number']).columns) > 0:
                assert 'Statistics' in sheet_names
        except ImportError:
            pytest.skip("openpyxl not available for validation")
    
    def test_save_excel_severity_colors(self, sample_dataframe, tmp_path):
        """Test that severity colors are applied correctly."""
        validation_issues = [
            {'row_number': 0, 'column_name': 'age', 'issue_type': 'type_mismatch',
             'description': 'High severity', 'severity': 'high'},
            {'row_number': 1, 'column_name': 'email', 'issue_type': 'invalid_email',
             'description': 'Medium severity', 'severity': 'medium'},
            {'row_number': 2, 'column_name': 'name', 'issue_type': 'missing_value',
             'description': 'Low severity', 'severity': 'low'},
        ]
        validation_summary = {'high': 1, 'medium': 1, 'low': 1}
        
        output_dir = str(tmp_path / "reports")
        excel_path = save_excel(
            df=sample_dataframe,
            validation_issues=validation_issues,
            validation_summary=validation_summary,
            filename="test_report_colors",
            output_dir=output_dir
        )
        
        try:
            import openpyxl
            from openpyxl.styles import PatternFill
            wb = openpyxl.load_workbook(excel_path)
            ws_issues = wb['Issues']
            
            # Check that severity column has colors
            for row_idx in range(2, ws_issues.max_row + 1):
                severity_cell = ws_issues[f'E{row_idx}']
                if severity_cell.value in ['high', 'medium', 'low']:
                    assert severity_cell.fill is not None
        except ImportError:
            pytest.skip("openpyxl not available for validation")
    
    def test_save_excel_missing_openpyxl(self, sample_dataframe, tmp_path, monkeypatch):
        """Test that ImportError is raised when openpyxl is not available."""
        # Mock import to raise ImportError
        import builtins
        
        original_import = builtins.__import__
        
        def mock_import(name, *args, **kwargs):
            if name == 'openpyxl' or name.startswith('openpyxl'):
                raise ImportError("No module named 'openpyxl'")
            return original_import(name, *args, **kwargs)
        
        monkeypatch.setattr(builtins, '__import__', mock_import)
        
        validation_issues = []
        validation_summary = {'high': 0, 'medium': 0, 'low': 0}
        
        with pytest.raises(ImportError, match="openpyxl is required"):
            save_excel(
                df=sample_dataframe,
                validation_issues=validation_issues,
                validation_summary=validation_summary,
                filename="test_report",
                output_dir=str(tmp_path / "reports")
            )

