
import os
import markdown
import pdfkit
from pdfkit.configuration import Configuration

# Compute absolute path to the HTML template
TEMPLATE_PATH = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "templates", "report_template.html")
)


def render_template(content: str, template_path: str) -> str:
    """Reads an HTML template and injects the Markdown HTML content."""
    with open(template_path, "r", encoding="utf-8") as f:
        template = f.read()
    return template.replace("{{content}}", content)


def save_markdown(report_md: str, filename: str, output_dir: str = "reports") -> str:
    """Saves raw Markdown content to a .md file."""
    os.makedirs(output_dir, exist_ok=True)
    md_path = os.path.join(output_dir, f"{filename}.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(report_md)
    return md_path


def save_html(report_md: str, filename: str, output_dir: str = "reports", visualizations: dict = None) -> str:
    """
    Converts Markdown to styled HTML using the template and saves it.
    
    Args:
        report_md: Markdown content
        filename: Output filename (without extension)
        output_dir: Output directory
        visualizations: Dictionary of base64-encoded chart images
        
    Returns:
        Path to created HTML file
    """
    os.makedirs(output_dir, exist_ok=True)
    html_body = markdown.markdown(report_md, extensions=['tables'])
    
    # Add visualizations if provided
    if visualizations:
        viz_html = "<div class='visualizations'>\n<h2>📊 Visualizations</h2>\n"
        
        if visualizations.get("missing_values"):
            viz_html += f"""
            <div class='chart-container'>
                <h3>Missing Values Count</h3>
                <img src="data:image/png;base64,{visualizations['missing_values']}" alt="Missing Values Chart" class="chart-img">
            </div>
            """
        
        if visualizations.get("missing_percentage"):
            viz_html += f"""
            <div class='chart-container'>
                <h3>Missing Values Percentage</h3>
                <img src="data:image/png;base64,{visualizations['missing_percentage']}" alt="Missing Percentage Chart" class="chart-img">
            </div>
            """
        
        if visualizations.get("issues_severity"):
            viz_html += f"""
            <div class='chart-container'>
                <h3>Issues by Severity</h3>
                <img src="data:image/png;base64,{visualizations['issues_severity']}" alt="Issues Severity Chart" class="chart-img">
            </div>
            """
        
        # Add numeric distributions
        if visualizations.get("numeric_distributions"):
            viz_html += "<div class='chart-container'><h3>Numeric Column Distributions</h3>"
            for col_name, chart_img in visualizations["numeric_distributions"].items():
                viz_html += f"""
                <div class='distribution-chart'>
                    <img src="data:image/png;base64,{chart_img}" alt="Distribution of {col_name}" class="chart-img">
                </div>
                """
            viz_html += "</div>"
        
        viz_html += "</div>"
        html_body = viz_html + html_body
    
    full_html = render_template(html_body, TEMPLATE_PATH)

    html_path = os.path.join(output_dir, f"{filename}.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(full_html)
    return html_path


def save_pdf(report_md: str, filename: str, output_dir: str = "reports") -> str:
    """Converts Markdown to styled HTML and renders it as PDF using wkhtmltopdf."""
    os.makedirs(output_dir, exist_ok=True)
    html_body = markdown.markdown(report_md, extensions=['tables'])
    full_html = render_template(html_body, TEMPLATE_PATH)

    temp_html_path = os.path.join(output_dir, f"{filename}.temp.html")
    with open(temp_html_path, "w", encoding="utf-8") as f:
        f.write(full_html)

    # Update this path if wkhtmltopdf is installed elsewhere
    config = Configuration(wkhtmltopdf=r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe")
    pdf_path = os.path.join(output_dir, f"{filename}.pdf")
    pdfkit.from_file(temp_html_path, pdf_path, configuration=config)

    os.remove(temp_html_path)
    return pdf_path


def save_excel(
    df: "pd.DataFrame",
    validation_issues: list,
    validation_summary: dict,
    ml_recommendations: dict = None,
    filename: str = "report",
    output_dir: str = "reports"
) -> str:
    """
    Export data quality report to Excel format with multiple sheets.
    
    Args:
        df: Input DataFrame
        validation_issues: List of validation issue dictionaries
        validation_summary: Validation summary dictionary
        ml_recommendations: Optional ML recommendations dictionary
        filename: Output filename (without extension)
        output_dir: Output directory
        
    Returns:
        Path to created Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from datetime import datetime
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
    
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, f"{filename}.xlsx")
    
    wb = Workbook()
    
    # Sheet 1: Overview
    ws_overview = wb.active
    ws_overview.title = "Overview"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Write overview data
    ws_overview['A1'] = "Data Quality Report"
    ws_overview['A1'].font = Font(bold=True, size=14)
    ws_overview['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_overview['A4'] = "Dataset Overview"
    ws_overview['A4'].font = Font(bold=True, size=12)
    ws_overview['A5'] = f"Rows: {df.shape[0]}"
    ws_overview['A6'] = f"Columns: {df.shape[1]}"
    ws_overview['A8'] = "Validation Summary"
    ws_overview['A8'].font = Font(bold=True, size=12)
    ws_overview['A9'] = f"Total Issues: {len(validation_issues)}"
    ws_overview['A10'] = f"High Severity: {validation_summary.get('high', 0)}"
    ws_overview['A11'] = f"Medium Severity: {validation_summary.get('medium', 0)}"
    ws_overview['A12'] = f"Low Severity: {validation_summary.get('low', 0)}"
    
    if ml_recommendations:
        ws_overview['A14'] = "ML Readiness"
        ws_overview['A14'].font = Font(bold=True, size=12)
        ws_overview['A15'] = f"Score: {ml_recommendations.get('readiness_score', 0)}/100"
        ws_overview['A16'] = f"Level: {ml_recommendations.get('readiness_level', 'Unknown')}"
    
    # Sheet 2: Missing Values
    ws_missing = wb.create_sheet("Missing Values")
    ws_missing['A1'] = "Column"
    ws_missing['B1'] = "Missing Count"
    ws_missing['C1'] = "Missing Percentage"
    
    for cell in ['A1', 'B1', 'C1']:
        ws_missing[cell].fill = header_fill
        ws_missing[cell].font = header_font
    
    missing_data = df.isnull().sum()
    missing_pct = (missing_data / len(df)) * 100
    
    row = 2
    for col in df.columns:
        if missing_data[col] > 0:
            ws_missing[f'A{row}'] = col
            ws_missing[f'B{row}'] = missing_data[col]
            ws_missing[f'C{row}'] = f"{missing_pct[col]:.2f}%"
            row += 1
    
    # Sheet 3: Issues
    ws_issues = wb.create_sheet("Issues")
    ws_issues['A1'] = "Row"
    ws_issues['B1'] = "Column"
    ws_issues['C1'] = "Issue Type"
    ws_issues['D1'] = "Description"
    ws_issues['E1'] = "Severity"
    
    for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
        ws_issues[cell].fill = header_fill
        ws_issues[cell].font = header_font
    
    row = 2
    for issue in validation_issues:
        ws_issues[f'A{row}'] = issue.get('row_number', 'N/A')
        ws_issues[f'B{row}'] = issue.get('column_name', 'N/A')
        ws_issues[f'C{row}'] = issue.get('issue_type', 'N/A')
        ws_issues[f'D{row}'] = issue.get('description', 'N/A')
        severity = issue.get('severity', 'medium')
        ws_issues[f'E{row}'] = severity
        
        # Color code by severity
        if severity == 'high':
            ws_issues[f'E{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        elif severity == 'medium':
            ws_issues[f'E{row}'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        else:
            ws_issues[f'E{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        row += 1
    
    # Sheet 4: Statistics
    ws_stats = wb.create_sheet("Statistics")
    numeric_df = df.select_dtypes(include=['number'])
    
    if len(numeric_df.columns) > 0:
        stats = numeric_df.describe()
        
        # Write headers
        ws_stats['A1'] = "Statistic"
        for idx, col in enumerate(stats.columns, start=2):
            ws_stats.cell(row=1, column=idx, value=col)
        
        for cell in range(1, len(stats.columns) + 2):
            ws_stats.cell(row=1, column=cell).fill = header_fill
            ws_stats.cell(row=1, column=cell).font = header_font
        
        # Write statistics
        row = 2
        for stat_name in stats.index:
            ws_stats.cell(row=row, column=1, value=stat_name)
            for col_idx, col_name in enumerate(stats.columns, start=2):
                ws_stats.cell(row=row, column=col_idx, value=stats.loc[stat_name, col_name])
            row += 1
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_path)
    return excel_path
